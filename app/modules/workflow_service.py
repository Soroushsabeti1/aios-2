"""
سرویس فلوهای کاری — ایجاد، اجرا، ذخیره خودکار.
"""
import json
from datetime import datetime, timezone
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession
from app.database.models.business import WorkFlow


async def create_workflow(session: AsyncSession, tenant_id: int,
                           name: str, trigger_type: str,
                           trigger_condition: dict,
                           steps: list[dict],
                           target_type: str = None,
                           target_id: int = None,
                           target_role: str = None,
                           max_retries: int = 3) -> str:
    flow = WorkFlow(
        tenant_id=tenant_id, name=name,
        trigger_type=trigger_type,
        trigger_condition=json.dumps(trigger_condition, ensure_ascii=False),
        steps_json=json.dumps(steps, ensure_ascii=False),
        target_type=target_type, target_id=target_id, target_role=target_role,
        max_retries=max_retries, is_active=True,
    )
    session.add(flow)
    await session.commit()
    return f"✅ فلو «{name}» ثبت شد."


async def list_workflows(session: AsyncSession, tenant_id: int,
                         detail: bool = False) -> str:
    flows = (await session.scalars(
        select(WorkFlow).where(WorkFlow.tenant_id == tenant_id)
        .order_by(WorkFlow.id.desc())
    )).all()
    if not flows:
        return "هیچ فلویی ثبت نشده."

    lines = [f"⚙️ فلوهای کاری ({len(flows)} فلو):"]
    for f in flows:
        status = "🟢 فعال" if f.is_active else "🔴 غیرفعال"
        cond = json.loads(f.trigger_condition) if f.trigger_condition else {}
        steps = json.loads(f.steps_json) if f.steps_json else []
        lines.append(f"\n[{f.id}] {f.name} — {status}")
        lines.append(f"   نوع: {f.trigger_type}")
        if cond.get("description"):
            lines.append(f"   شرط: {cond['description']}")
        if cond.get("event"):
            lines.append(f"   رویداد: {cond['event']}")
        if f.target_role:
            lines.append(f"   هدف: {f.target_role}")
        if steps:
            lines.append(f"   مراحل ({len(steps)}):")
            for i, s in enumerate(steps, 1):
                action = s.get("action", "")
                msg = s.get("message", "")
                target = s.get("target_role", s.get("target", ""))
                lines.append(f"   {i}. {action}" + (f" → {target}" if target else "") + (f": {msg[:50]}" if msg else ""))
        lines.append(f"   برای حذف: «فلو {f.id} رو حذف کن»")
    return "\n".join(lines)


async def get_workflow_detail(session: AsyncSession, tenant_id: int,
                               flow_id: int) -> str:
    flow = await session.get(WorkFlow, flow_id)
    if not flow or flow.tenant_id != tenant_id:
        return "⚠️ فلو پیدا نشد."
    return await list_workflows(session, tenant_id, detail=True)


async def toggle_workflow(session: AsyncSession, tenant_id: int,
                           flow_id: int, active: bool) -> str:
    flow = await session.get(WorkFlow, flow_id)
    if not flow or flow.tenant_id != tenant_id:
        return "⚠️ فلو پیدا نشد."
    flow.is_active = active
    await session.commit()
    return f"✅ فلو «{flow.name}» {'فعال' if active else 'غیرفعال'} شد."


async def delete_workflow(session: AsyncSession, tenant_id: int, flow_id: int) -> str:
    flow = await session.get(WorkFlow, flow_id)
    if not flow or flow.tenant_id != tenant_id:
        return "⚠️ فلو پیدا نشد."
    name = flow.name
    await session.delete(flow)
    await session.commit()
    return f"🗑 فلو «{name}» حذف شد."


async def export_workflows_excel(session: AsyncSession, tenant_id: int):
    """خروجی اکسل فلوها."""
    import io
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill

    flows = (await session.scalars(
        select(WorkFlow).where(WorkFlow.tenant_id == tenant_id)
    )).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "فلوها"
    ws.sheet_view.rightToLeft = True

    headers = ["شناسه", "نام فلو", "نوع شرط", "توضیح شرط", "مراحل", "هدف", "وضعیت"]
    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=i, value=h)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="2B5F9E")
        cell.font = Font(bold=True, color="FFFFFF")

    for row_idx, flow in enumerate(flows, 2):
        cond = json.loads(flow.trigger_condition) if flow.trigger_condition else {}
        steps = json.loads(flow.steps_json) if flow.steps_json else []
        steps_desc = " ← ".join([s.get("action", "") for s in steps[:3]])
        ws.cell(row=row_idx, column=1, value=flow.id)
        ws.cell(row=row_idx, column=2, value=flow.name)
        ws.cell(row=row_idx, column=3, value=flow.trigger_type)
        ws.cell(row=row_idx, column=4, value=cond.get("description", ""))
        ws.cell(row=row_idx, column=5, value=steps_desc)
        ws.cell(row=row_idx, column=6, value=flow.target_role or flow.target_type or "همه")
        ws.cell(row=row_idx, column=7, value="فعال" if flow.is_active else "غیرفعال")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf, "فلوهای_کاری.xlsx"
