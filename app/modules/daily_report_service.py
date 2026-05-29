"""سرویس گزارش روزانه کارمند — بهره‌وری، حضور، تسک‌ها."""
import io
import json
from datetime import datetime, timezone, date
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession
from app.database.models.business import DailyReport, Person
from app.utils.normalizer import format_amount


def calc_productivity(tasks: list[dict], total_work_minutes: int) -> float:
    """
    فرمول بهره‌وری: (زمان مفید / کل تایم کاری) × ۱۰۰
    زمان مفید = مجموع دقایق تسک‌های تکمیل‌شده
    """
    if not total_work_minutes:
        return 0.0
    done_minutes = sum(
        t.get("duration_minutes", 0)
        for t in tasks
        if t.get("status") in ("done", "approved")
    )
    return min(100.0, round((done_minutes / total_work_minutes) * 100, 1))


def productivity_icon(pct: float) -> str:
    if pct >= 80:
        return "🟢"
    if pct >= 60:
        return "🟡"
    return "🔴"


async def submit_daily_report(session: AsyncSession, tenant_id: int,
                               person_telegram_id: int,
                               check_in_str: str = None,
                               check_out_str: str = None,
                               break_minutes: int = 0,
                               overtime_hours: float = 0,
                               night_hours: float = 0,
                               holiday_work: bool = False,
                               friday_work: bool = False,
                               tasks_data: list[dict] = None,
                               submitted_by: str = "employee") -> str:
    person = await session.scalar(
        select(Person).where(Person.telegram_id == person_telegram_id,
                              Person.is_active == True)
    )
    if not person:
        return "⚠️ حساب پیدا نشد."

    today = date.today()

    # چک گزارش تکراری
    existing = await session.scalar(
        select(DailyReport).where(
            DailyReport.person_id == person.id,
            DailyReport.report_date >= datetime.combine(today, datetime.min.time()).replace(tzinfo=timezone.utc),
        )
    )
    if existing and existing.is_confirmed:
        return "⚠️ گزارش امروزت قبلاً ثبت و قفل شده."

    # محاسبه زمان کاری
    check_in = check_out = None
    total_work_minutes = 0
    if check_in_str and check_out_str:
        try:
            ci = datetime.strptime(check_in_str, "%H:%M").replace(
                year=today.year, month=today.month, day=today.day, tzinfo=timezone.utc)
            co = datetime.strptime(check_out_str, "%H:%M").replace(
                year=today.year, month=today.month, day=today.day, tzinfo=timezone.utc)
            check_in, check_out = ci, co
            total_work_minutes = max(0, int((co - ci).total_seconds() / 60) - break_minutes)
        except Exception:
            pass

    # بهره‌وری
    productivity = calc_productivity(tasks_data or [], total_work_minutes) if total_work_minutes else None

    report = existing or DailyReport(
        tenant_id=tenant_id, person_id=person.id,
        report_date=datetime.now(timezone.utc),
    )
    report.check_in = check_in
    report.check_out = check_out
    report.break_minutes = break_minutes
    report.overtime_hours = overtime_hours
    report.night_hours = night_hours
    report.holiday_work = holiday_work
    report.friday_work = friday_work
    report.tasks_json = json.dumps(tasks_data or [], ensure_ascii=False)
    report.productivity = productivity
    report.submitted_by = submitted_by

    if not existing:
        session.add(report)
    await session.flush()

    # اگه تأیید نشده، نشون بده
    if not report.is_confirmed:
        tasks_summary = ""
        if tasks_data:
            tasks_summary = "\n\nتسک‌های امروز:\n" + "\n".join(
                f"• {t.get('title','')} — {t.get('status','')} — {t.get('duration_minutes',0)} دقیقه"
                for t in tasks_data[:5]
            )
        prod_str = f"{productivity_icon(productivity)} {productivity}%" if productivity is not None else "—"
        preview = (
            f"📋 پیش‌نمایش گزارش امروز:\n"
            f"ورود: {check_in_str or '—'} | خروج: {check_out_str or '—'}\n"
            f"استراحت: {break_minutes} دقیقه | اضافه‌کاری: {overtime_hours} ساعت\n"
            f"بهره‌وری: {prod_str}{tasks_summary}\n\n"
            f"ثبت می‌کنم؟"
        )
        await session.rollback()
        return f"PREVIEW|{preview}"

    await session.commit()
    return "✅ گزارش ثبت شد و قفل شد."


async def confirm_report(session: AsyncSession, tenant_id: int,
                          person_telegram_id: int) -> str:
    person = await session.scalar(
        select(Person).where(Person.telegram_id == person_telegram_id)
    )
    if not person:
        return "⚠️ حساب پیدا نشد."
    today = date.today()
    report = await session.scalar(
        select(DailyReport).where(
            DailyReport.person_id == person.id,
            DailyReport.report_date >= datetime.combine(today, datetime.min.time()).replace(tzinfo=timezone.utc),
        )
    )
    if not report:
        return "⚠️ گزارشی برای تأیید پیدا نشد."
    report.is_confirmed = True
    await session.commit()
    return "✅ گزارش قفل شد."


async def get_daily_report(session: AsyncSession, tenant_id: int,
                            person_id: int, report_date: date = None) -> str:
    """دریافت گزارش یک کارمند برای یک روز."""
    target = report_date or date.today()
    report = await session.scalar(
        select(DailyReport).where(
            DailyReport.person_id == person_id,
            DailyReport.report_date >= datetime.combine(target, datetime.min.time()).replace(tzinfo=timezone.utc),
        )
    )
    if not report:
        return "گزارشی برای این روز پیدا نشد."

    tasks = json.loads(report.tasks_json) if report.tasks_json else []
    prod = f"{productivity_icon(report.productivity)} {report.productivity}%" if report.productivity is not None else "—"

    lines = [
        f"📋 گزارش روزانه:",
        f"ورود: {report.check_in.strftime('%H:%M') if report.check_in else '—'} | "
        f"خروج: {report.check_out.strftime('%H:%M') if report.check_out else '—'}",
        f"استراحت: {report.break_minutes} دقیقه",
        f"اضافه‌کاری: {report.overtime_hours} ساعت | شب‌کاری: {report.night_hours} ساعت",
        f"بهره‌وری: {prod}",
    ]
    if tasks:
        lines.append(f"\nتسک‌ها ({len(tasks)}):")
        for t in tasks:
            lines.append(f"• {t.get('title','')} [{t.get('type','')}] — {t.get('status','')} — {t.get('duration_minutes',0)} دقیقه")
    if report.positives:
        lines.append(f"\n✅ نکات مثبت: {report.positives}")
    if report.negatives:
        lines.append(f"⚠️ نکات منفی: {report.negatives}")
    lines.append(f"\nوضعیت: {'✅ قفل‌شده' if report.is_confirmed else '⏳ پیش‌نویس'}")
    return "\n".join(lines)


async def end_of_day_report(session: AsyncSession, tenant_id: int) -> tuple[str, bytes | None]:
    """گزارش پایان روز — همه کارمندان."""
    today = date.today()
    reports = (await session.scalars(
        select(DailyReport).where(
            DailyReport.tenant_id == tenant_id,
            DailyReport.report_date >= datetime.combine(today, datetime.min.time()).replace(tzinfo=timezone.utc),
        )
    )).all()

    if not reports:
        return "هیچ گزارشی برای امروز نداریم.", None

    lines = [f"📊 گزارش پایان روز — {today}:"]
    total_prod = []

    for r in reports:
        person = await session.get(Person, r.person_id)
        name = person.full_name if person else "—"
        prod = r.productivity or 0
        total_prod.append(prod)
        icon = productivity_icon(prod)
        lines.append(f"{icon} {name}: {prod}% بهره‌وری | {r.no_response_count} عدم پاسخ")

    if total_prod:
        avg = round(sum(total_prod) / len(total_prod), 1)
        lines.append(f"\n{productivity_icon(avg)} میانگین مجموعه: {avg}%")

    return "\n".join(lines), None


async def export_daily_report_excel(session: AsyncSession,
                                     report: DailyReport,
                                     person_name: str) -> io.BytesIO:
    """اکسل گزارش روزانه یک کارمند."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "گزارش روزانه"
    ws.sheet_view.rightToLeft = True

    # هدر
    ws["A1"] = "گزارش روزانه"
    ws["A2"] = f"کارمند: {person_name}"
    ws["A3"] = f"تاریخ: {report.report_date.date() if report.report_date else '—'}"
    ws["A4"] = f"ورود: {report.check_in.strftime('%H:%M') if report.check_in else '—'}"
    ws["B4"] = f"خروج: {report.check_out.strftime('%H:%M') if report.check_out else '—'}"
    ws["A5"] = f"استراحت: {report.break_minutes} دقیقه"
    ws["B5"] = f"اضافه‌کاری: {report.overtime_hours} ساعت"
    ws["A6"] = f"بهره‌وری: {report.productivity or 0}%"

    # تسک‌ها
    ws["A8"] = "تسک‌ها"
    ws["A8"].font = Font(bold=True)
    headers = ["عنوان", "نوع", "وضعیت", "مدت (دقیقه)", "خروجی", "تحویل به"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=9, column=i, value=h).font = Font(bold=True)

    tasks = json.loads(report.tasks_json) if report.tasks_json else []
    for r_idx, t in enumerate(tasks, 10):
        ws.cell(row=r_idx, column=1, value=t.get("title", ""))
        ws.cell(row=r_idx, column=2, value=t.get("type", ""))
        ws.cell(row=r_idx, column=3, value=t.get("status", ""))
        ws.cell(row=r_idx, column=4, value=t.get("duration_minutes", 0))
        ws.cell(row=r_idx, column=5, value=t.get("output", ""))
        ws.cell(row=r_idx, column=6, value=t.get("delivered_to", ""))

    if report.positives:
        r = len(tasks) + 12
        ws.cell(row=r, column=1, value="نکات مثبت:")
        ws.cell(row=r, column=2, value=report.positives)
    if report.negatives:
        r = len(tasks) + 13
        ws.cell(row=r, column=1, value="نکات منفی:")
        ws.cell(row=r, column=2, value=report.negatives)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
