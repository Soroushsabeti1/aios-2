"""
اکسل فاکتور حرفه‌ای — با فرمول، لوگو، محل مهر و امضا.
قابل استفاده دستی توسط کارفرما.
"""
import io
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XlImage
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession
from app.database.models.business import Invoice, InvoiceItem
from app.database.models.tenant import Tenant
from app.utils.normalizer import format_amount
from app.utils.jalali import to_jalali_str

HEADER_FILL = PatternFill("solid", start_color="3B82F6")
HEADER_FONT = Font(bold=True, color="FFFFFF", name="Arial", size=11)
TITLE_FONT = Font(bold=True, name="Arial", size=14)
CELL_FONT = Font(name="Arial", size=10)
BOLD_FONT = Font(bold=True, name="Arial", size=11)
SMALL_FONT = Font(name="Arial", size=9, italic=True, color="6B7280")
CENTER = Alignment(horizontal="center", vertical="center")
RIGHT = Alignment(horizontal="right", vertical="center")
BORDER = Border(*(Side(style="thin", color="E5E7EB"),) * 4)
DASHED = Border(
    top=Side(style="dashed", color="CCCCCC"),
    bottom=Side(style="dashed", color="CCCCCC"),
    left=Side(style="dashed", color="CCCCCC"),
    right=Side(style="dashed", color="CCCCCC"),
)
STATUS_COLORS = {
    "draft": PatternFill("solid", start_color="FEF3C7"),
    "confirmed": PatternFill("solid", start_color="DBEAFE"),
    "paid": PatternFill("solid", start_color="DCFCE7"),
    "cancelled": PatternFill("solid", start_color="FEE2E2"),
}
STATUS_LABELS = {"draft": "پیش‌فاکتور", "confirmed": "تأیید‌شده", "paid": "پرداخت‌شده", "cancelled": "کنسل"}


async def make_invoice_excel(session: AsyncSession, tenant_id: int,
                             invoice_display_id: str) -> tuple[io.BytesIO | None, str | None, str | None]:
    """ساخت اکسل یک فاکتور."""
    invoice = await session.scalar(
        select(Invoice).where(Invoice.tenant_id == tenant_id, Invoice.display_id == invoice_display_id)
    )
    if not invoice:
        return None, None, f"فاکتور {invoice_display_id} پیدا نشد."

    tenant = await session.get(Tenant, tenant_id)
    items = (await session.scalars(
        select(InvoiceItem).where(InvoiceItem.invoice_id == invoice.id)
    )).all()

    wb = Workbook()
    ws = wb.active
    ws.title = f"فاکتور {invoice.display_id}"
    ws.sheet_view.rightToLeft = True
    ws.page_setup.orientation = "portrait"
    ws.print_options.horizontalCentered = True

    # عرض ستون‌ها: A(ردیف) B(شناسه) C(کالا) D(تعداد) E(فی) F(جمع)
    widths = [8, 12, 28, 10, 18, 18]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    row = 1

    # ─── لوگو + اطلاعات فروشگاه ───
    if tenant and tenant.logo:
        try:
            logo_stream = io.BytesIO(tenant.logo)
            img = XlImage(logo_stream)
            img.width = 60
            img.height = 60
            ws.add_image(img, "F1")
        except Exception:
            pass

    ws.merge_cells("A1:E1")
    c = ws.cell(row=1, column=1, value=tenant.name if tenant else "فروشگاه")
    c.font = TITLE_FONT
    c.alignment = RIGHT
    row = 2

    if tenant:
        info_parts = []
        if tenant.phone:
            info_parts.append(f"تلفن: {tenant.phone}")
        if tenant.address:
            info_parts.append(f"آدرس: {tenant.address}")
        if info_parts:
            ws.merge_cells(f"A{row}:F{row}")
            ws.cell(row=row, column=1, value=" | ".join(info_parts)).font = SMALL_FONT
            ws.cell(row=row, column=1).alignment = RIGHT
            row += 1

    row += 1

    # ─── شماره فاکتور + وضعیت + تاریخ ───
    ws.merge_cells(f"A{row}:C{row}")
    ws.cell(row=row, column=1, value=f"شماره فاکتور: {invoice.display_id}").font = BOLD_FONT
    ws.cell(row=row, column=1).alignment = RIGHT
    ws.merge_cells(f"D{row}:F{row}")
    status_cell = ws.cell(row=row, column=4, value=STATUS_LABELS.get(invoice.status, invoice.status))
    status_cell.font = BOLD_FONT
    status_cell.fill = STATUS_COLORS.get(invoice.status, PatternFill())
    status_cell.alignment = CENTER
    row += 1

    ws.merge_cells(f"A{row}:C{row}")
    ws.cell(row=row, column=1, value=f"مشتری: {invoice.customer_name or '—'}").font = CELL_FONT
    ws.cell(row=row, column=1).alignment = RIGHT
    ws.merge_cells(f"D{row}:F{row}")
    inv_date = to_jalali_str(invoice.invoice_date) if invoice.invoice_date else "—"
    ws.cell(row=row, column=4, value=f"تاریخ: {inv_date}").font = CELL_FONT
    ws.cell(row=row, column=4).alignment = CENTER
    row += 2

    # ─── هدر جدول ───
    headers = ["ردیف", "شناسه", "نام کالا", "تعداد", "فی (تومان)", "جمع (تومان)"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=col, value=h)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = CENTER
        c.border = BORDER
    row += 1

    # ─── آیتم‌ها ───
    data_start = row
    for i, item in enumerate(items, 1):
        ws.cell(row=row, column=1, value=i).font = CELL_FONT
        ws.cell(row=row, column=1).alignment = CENTER
        ws.cell(row=row, column=1).border = BORDER
        ws.cell(row=row, column=2, value="").font = CELL_FONT  # شناسه کالا
        ws.cell(row=row, column=2).alignment = CENTER
        ws.cell(row=row, column=2).border = BORDER
        ws.cell(row=row, column=3, value=item.product_name).font = CELL_FONT
        ws.cell(row=row, column=3).alignment = RIGHT
        ws.cell(row=row, column=3).border = BORDER
        ws.cell(row=row, column=4, value=item.quantity).font = CELL_FONT
        ws.cell(row=row, column=4).alignment = CENTER
        ws.cell(row=row, column=4).border = BORDER
        ws.cell(row=row, column=5, value=float(item.unit_price)).font = CELL_FONT
        ws.cell(row=row, column=5).number_format = "#,##0"
        ws.cell(row=row, column=5).alignment = CENTER
        ws.cell(row=row, column=5).border = BORDER
        # فرمول جمع = تعداد × فی
        ws.cell(row=row, column=6).font = BOLD_FONT
        ws.cell(row=row, column=6).number_format = "#,##0"
        ws.cell(row=row, column=6).alignment = CENTER
        ws.cell(row=row, column=6).border = BORDER
        ws.cell(row=row, column=6, value=f"=D{row}*E{row}")
        row += 1

    # ── ردیف‌های خالی اضافه برای پر کردن دستی ──
    for _ in range(5):
        for col in range(1, 7):
            c = ws.cell(row=row, column=col, value="")
            c.border = BORDER
            c.font = CELL_FONT
        ws.cell(row=row, column=6, value=f"=D{row}*E{row}")
        ws.cell(row=row, column=6).number_format = "#,##0"
        row += 1

    data_end = row - 1
    row += 1

    # ─── جمع‌ها با فرمول ───
    ws.merge_cells(f"A{row}:D{row}")
    ws.cell(row=row, column=1, value="جمع کل:").font = BOLD_FONT
    ws.cell(row=row, column=1).alignment = RIGHT
    ws.merge_cells(f"E{row}:F{row}")
    ws.cell(row=row, column=5, value=f"=SUM(F{data_start}:F{data_end})").font = BOLD_FONT
    ws.cell(row=row, column=5).number_format = "#,##0"
    ws.cell(row=row, column=5).alignment = CENTER
    total_row = row
    row += 1

    ws.merge_cells(f"A{row}:D{row}")
    ws.cell(row=row, column=1, value="تخفیف:").font = CELL_FONT
    ws.cell(row=row, column=1).alignment = RIGHT
    ws.merge_cells(f"E{row}:F{row}")
    ws.cell(row=row, column=5, value=float(invoice.discount or 0)).font = CELL_FONT
    ws.cell(row=row, column=5).number_format = "#,##0"
    ws.cell(row=row, column=5).alignment = CENTER
    discount_row = row
    row += 1

    tax_pct = 9
    if tenant and tenant.default_tax_percent:
        tax_pct = tenant.default_tax_percent
    ws.merge_cells(f"A{row}:D{row}")
    ws.cell(row=row, column=1, value=f"مالیات ({tax_pct}٪):").font = CELL_FONT
    ws.cell(row=row, column=1).alignment = RIGHT
    ws.merge_cells(f"E{row}:F{row}")
    ws.cell(row=row, column=5, value=f"=E{total_row}*{tax_pct/100}").font = CELL_FONT
    ws.cell(row=row, column=5).number_format = "#,##0"
    ws.cell(row=row, column=5).alignment = CENTER
    tax_row = row
    row += 1

    ws.merge_cells(f"A{row}:D{row}")
    ws.cell(row=row, column=1, value="💰 مبلغ نهایی:").font = Font(bold=True, name="Arial", size=13)
    ws.cell(row=row, column=1).alignment = RIGHT
    ws.merge_cells(f"E{row}:F{row}")
    ws.cell(row=row, column=5, value=f"=E{total_row}-E{discount_row}+E{tax_row}").font = Font(bold=True, name="Arial", size=13)
    ws.cell(row=row, column=5).number_format = "#,##0"
    ws.cell(row=row, column=5).alignment = CENTER
    row += 2

    # ─── اطلاعات پرداخت ───
    if tenant and (tenant.card_number or tenant.sheba):
        ws.merge_cells(f"A{row}:F{row}")
        parts = []
        if tenant.card_number:
            parts.append(f"شماره کارت: {tenant.card_number}")
        if tenant.sheba:
            parts.append(f"شبا: {tenant.sheba}")
        if tenant.account_holder:
            parts.append(f"به نام: {tenant.account_holder}")
        ws.cell(row=row, column=1, value=" | ".join(parts)).font = CELL_FONT
        ws.cell(row=row, column=1).alignment = RIGHT
        row += 2

    # ─── مهر و امضا ───
    ws.merge_cells(f"A{row}:C{row}")
    ws.cell(row=row, column=1, value="مهر و امضای فروشنده").font = SMALL_FONT
    ws.cell(row=row, column=1).alignment = CENTER
    ws.merge_cells(f"D{row}:F{row}")
    ws.cell(row=row, column=4, value="امضای خریدار").font = SMALL_FONT
    ws.cell(row=row, column=4).alignment = CENTER
    row += 1

    # فضای خالی برای مهر
    for _ in range(4):
        ws.merge_cells(f"A{row}:C{row}")
        ws.cell(row=row, column=1).border = DASHED
        ws.merge_cells(f"D{row}:F{row}")
        ws.cell(row=row, column=4).border = DASHED
        row += 1

    row += 1
    ws.merge_cells(f"A{row}:F{row}")
    ws.cell(row=row, column=1,
            value="⚠️ این فاکتور پس از مهر و امضا معتبر است.").font = SMALL_FONT
    ws.cell(row=row, column=1).alignment = CENTER

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"فاکتور_{invoice.display_id}_{invoice.customer_name or ''}.xlsx"
    return buf, fname, None


async def make_blank_invoice_excel(session: AsyncSession, tenant_id: int) -> tuple[io.BytesIO, str]:
    """اکسل فاکتور خالی با فرمول — برای استفاده دستی."""
    tenant = await session.get(Tenant, tenant_id)

    wb = Workbook()
    ws = wb.active
    ws.title = "فاکتور"
    ws.sheet_view.rightToLeft = True

    widths = [8, 12, 28, 10, 18, 18]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    row = 1
    # لوگو
    if tenant and tenant.logo:
        try:
            img = XlImage(io.BytesIO(tenant.logo))
            img.width = 60
            img.height = 60
            ws.add_image(img, "F1")
        except Exception:
            pass

    ws.merge_cells("A1:E1")
    ws.cell(row=1, column=1, value=tenant.name if tenant else "نام فروشگاه").font = TITLE_FONT
    ws.cell(row=1, column=1).alignment = RIGHT
    row = 2

    if tenant:
        info = []
        if tenant.phone:
            info.append(f"تلفن: {tenant.phone}")
        if tenant.address:
            info.append(f"آدرس: {tenant.address}")
        if info:
            ws.merge_cells(f"A{row}:F{row}")
            ws.cell(row=row, column=1, value=" | ".join(info)).font = SMALL_FONT
            ws.cell(row=row, column=1).alignment = RIGHT
            row += 1

    row += 1

    # شماره + تاریخ + مشتری (خالی)
    ws.merge_cells(f"A{row}:C{row}")
    ws.cell(row=row, column=1, value="شماره فاکتور:").font = BOLD_FONT
    ws.merge_cells(f"D{row}:F{row}")
    ws.cell(row=row, column=4, value="تاریخ:").font = BOLD_FONT
    row += 1
    ws.merge_cells(f"A{row}:F{row}")
    ws.cell(row=row, column=1, value="مشتری:").font = BOLD_FONT
    row += 2

    # هدر
    headers = ["ردیف", "شناسه", "نام کالا", "تعداد", "فی (تومان)", "جمع (تومان)"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=col, value=h)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = CENTER
        c.border = BORDER
    row += 1

    # ۲۰ ردیف خالی با فرمول
    data_start = row
    for i in range(1, 21):
        ws.cell(row=row, column=1, value=i).font = CELL_FONT
        ws.cell(row=row, column=1).alignment = CENTER
        for col in range(1, 7):
            ws.cell(row=row, column=col).border = BORDER
        ws.cell(row=row, column=6, value=f"=D{row}*E{row}")
        ws.cell(row=row, column=6).number_format = "#,##0"
        ws.cell(row=row, column=5).number_format = "#,##0"
        row += 1
    data_end = row - 1
    row += 1

    # جمع‌ها
    tax_pct = tenant.default_tax_percent if tenant else 9
    ws.merge_cells(f"A{row}:D{row}")
    ws.cell(row=row, column=1, value="جمع کل:").font = BOLD_FONT
    ws.cell(row=row, column=1).alignment = RIGHT
    ws.merge_cells(f"E{row}:F{row}")
    ws.cell(row=row, column=5, value=f"=SUM(F{data_start}:F{data_end})").font = BOLD_FONT
    ws.cell(row=row, column=5).number_format = "#,##0"
    ws.cell(row=row, column=5).alignment = CENTER
    total_row = row
    row += 1

    ws.merge_cells(f"A{row}:D{row}")
    ws.cell(row=row, column=1, value="تخفیف:").font = CELL_FONT
    ws.cell(row=row, column=1).alignment = RIGHT
    ws.merge_cells(f"E{row}:F{row}")
    ws.cell(row=row, column=5, value=0).number_format = "#,##0"
    ws.cell(row=row, column=5).alignment = CENTER
    discount_row = row
    row += 1

    ws.merge_cells(f"A{row}:D{row}")
    ws.cell(row=row, column=1, value=f"مالیات ({tax_pct}٪):").font = CELL_FONT
    ws.cell(row=row, column=1).alignment = RIGHT
    ws.merge_cells(f"E{row}:F{row}")
    ws.cell(row=row, column=5, value=f"=E{total_row}*{tax_pct/100}").number_format = "#,##0"
    ws.cell(row=row, column=5).alignment = CENTER
    tax_row = row
    row += 1

    ws.merge_cells(f"A{row}:D{row}")
    ws.cell(row=row, column=1, value="مبلغ نهایی:").font = Font(bold=True, name="Arial", size=13)
    ws.cell(row=row, column=1).alignment = RIGHT
    ws.merge_cells(f"E{row}:F{row}")
    ws.cell(row=row, column=5, value=f"=E{total_row}-E{discount_row}+E{tax_row}").font = Font(bold=True, name="Arial", size=13)
    ws.cell(row=row, column=5).number_format = "#,##0"
    ws.cell(row=row, column=5).alignment = CENTER
    row += 2

    # پرداخت
    if tenant and (tenant.card_number or tenant.sheba):
        ws.merge_cells(f"A{row}:F{row}")
        parts = []
        if tenant.card_number:
            parts.append(f"شماره کارت: {tenant.card_number}")
        if tenant.sheba:
            parts.append(f"شبا: {tenant.sheba}")
        if tenant.account_holder:
            parts.append(f"به نام: {tenant.account_holder}")
        ws.cell(row=row, column=1, value=" | ".join(parts)).font = CELL_FONT
        ws.cell(row=row, column=1).alignment = RIGHT
        row += 2

    # مهر و امضا
    ws.merge_cells(f"A{row}:C{row}")
    ws.cell(row=row, column=1, value="مهر و امضای فروشنده").font = SMALL_FONT
    ws.cell(row=row, column=1).alignment = CENTER
    ws.merge_cells(f"D{row}:F{row}")
    ws.cell(row=row, column=4, value="امضای خریدار").font = SMALL_FONT
    ws.cell(row=row, column=4).alignment = CENTER
    row += 1
    for _ in range(4):
        ws.merge_cells(f"A{row}:C{row}")
        ws.cell(row=row, column=1).border = DASHED
        ws.merge_cells(f"D{row}:F{row}")
        ws.cell(row=row, column=4).border = DASHED
        row += 1

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf, "فاکتور_خالی.xlsx"
