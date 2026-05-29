"""
تولید PDF فیش تصفیه حساب.

روش: اکسل اصلی → جایگزینی مقادیر → فارسی‌سازی → LibreOffice → PDF
ظاهر ۱۰۰٪ مشابه اکسل اصلی.
"""
from __future__ import annotations
import io
import os
import re
import shutil
import subprocess
import tempfile

import openpyxl
from openpyxl.styles import Font
from openpyxl.worksheet.page import PageMargins
from openpyxl.worksheet.properties import WorksheetProperties, PageSetupProperties

from app.modules.reports.settlement_engine import SettlementResult

# مسیر اکسل template — کنار همین فایل
_TEMPLATE_DIR = os.path.dirname(os.path.abspath(__file__))
_TEMPLATE_PATH = os.path.join(_TEMPLATE_DIR, "settlement_template.xlsm")


def _to_fa(value) -> str:
    persian = '۰۱۲۳۴۵۶۷۸۹'
    return ''.join(persian[int(c)] if c.isdigit() else c for c in str(value))


def _fa_money(amount: float) -> str:
    return _to_fa(f"{int(round(amount)):,}")


def _fa_int(n) -> str:
    return _to_fa(str(int(n)))


def _fa_float1(n) -> str:
    if n == int(n):
        return _to_fa(str(int(n)))
    return _to_fa(f"{n:.1f}")


def _set_btitr(cell):
    old = cell.font
    cell.font = Font(
        name='B Titr', size=old.size, bold=old.bold,
        italic=old.italic, underline=old.underline, color=old.color,
    )


def generate_settlement_pdf(result: SettlementResult) -> tuple[io.BytesIO, str]:
    """
    تولید PDF فیش تصفیه حساب.
    Returns: (BytesIO حاوی PDF, نام فایل)
    """
    inp = result.inp
    r = result

    if not os.path.exists(_TEMPLATE_PATH):
        raise FileNotFoundError(
            f"فایل template پیدا نشد: {_TEMPLATE_PATH}\n"
            "فایل settlement_template.xlsm را کنار settlement_pdf.py بگذارید."
        )

    with tempfile.TemporaryDirectory() as tmpdir:
        # کپی template
        xlsx_path = os.path.join(tmpdir, 'settlement.xlsm')
        shutil.copy2(_TEMPLATE_PATH, xlsx_path)

        wb = openpyxl.load_workbook(xlsx_path, keep_vba=False)
        ws = wb['t,sahih(ساعت)']

        months = 1
        if inp.month_end != inp.month_start:
            months = inp.month_end - inp.month_start + 1

        # ═══════════════════════════════════════
        # ۱. جایگزینی مقادیر
        # ═══════════════════════════════════════

        # هدر
        ws['A7'] = inp.employee_name
        ws['I7'] = inp.employer_name
        ws['F5'] = inp.year
        ws['E9'] = 'تمــام وقت' if inp.work_type in ('تمام وقت', '') else 'نیــمه وقت'

        # B4: کد پرسنلی-ماه‌سال (مثال: 245158-0409)
        code_part = inp.employee_code or (inp.national_id[-6:] if inp.national_id else '')
        code_str = f"{code_part}-{inp.month_end:02d}{str(inp.year)[-2:]}" if code_part else ''
        ws['B4'] = code_str

        # حذف header صفحه (حاوی "g")
        ws.oddHeader.center.text = ''
        ws.oddHeader.left.text = ''
        ws.oddHeader.right.text = ''

        # پاک‌سازی ردیف ۱ و ۲ کامل (فرمول‌های خام و g)
        for row in range(1, 3):
            for col in range(1, 50):
                cell = ws.cell(row=row, column=col)
                if not isinstance(cell, openpyxl.cell.cell.MergedCell):
                    cell.value = None

        # پاک‌سازی سلول‌های فرمول‌دار خارج از جدول اصلی
        formula_cleanup = [
            'U3','P4','U4',
            'V3','V4','V5','V6','W3','W5','W6','W7',
            'Z4','Z5','Z6','Z7',
            'AC3','AC4','AC5','AC6','AC7','AC8','AC9','AC10','AC11',
            'W14','W16','W17','W19',
            'X2','X3','X5','X6','X12','X14','X16','X17','X19','X27','X28',
            'Y5','Y6','Y7','Y8','Y11',
            'W24','W25','W27','W28','W31',
            'U36','X36','U37','X37','Y37',
            'Z14','Z15','Z21','Z22','Z25','Z31',
            'X25','X26','Y17','Y19','Y20','Y21','Y22','Y23','Y25',
            'X23','W47','W48','Y47','Y48','Y49',
            'X34','X35','X39','X40',
            'AD3','AD4','AD5','AD6','AD7','AD8','AD9','AD10','AD11',
            'AA9','AA10','AA11',
            'V12','V20','V24','V27','V29',
            'Z26','Z48','Z49',
            'W39','W40','X39','X40',
            'Y12','Y13',
        ]
        for addr in formula_cleanup:
            ws[addr] = None

        # ردیف‌های ۳-۵۰ ستون‌های AF-AQ هم پاک (خارج print area)
        for row in range(1, 51):
            for col in range(32, 50):  # AF=32
                cell = ws.cell(row=row, column=col)
                if not isinstance(cell, openpyxl.cell.cell.MergedCell):
                    cell.value = None

        # تاریخ
        ws['D10'] = inp.year
        ws['F10'] = f"{inp.month_end:02d}"
        ws['H10'] = f"{inp.day_end:02d}"
        ws['N10'] = inp.year
        ws['P10'] = f"{inp.month_start:02d}"
        ws['R10'] = f"{inp.day_start:02d}"
        ws['K11'] = months
        ws['B12'] = inp.year
        ws['K12'] = inp.year

        # جدول — ستون C (مقادیر چپ)
        ws['C14'] = inp.work_hours
        ws['C15'] = r.hourly_wage
        ws['C16'] = r.housing_daily
        ws['C17'] = r.grocery_daily
        ws['C18'] = inp.marital_status or 'مجرد'
        ws['C21'] = inp.overtime_hours
        ws['C22'] = inp.friday_days
        ws['C23'] = inp.night_hours
        ws['C24'] = inp.holiday_days
        ws['C26'] = r.severance_daily
        ws['C27'] = r.bonus_daily

        # جدول — ستون L (مقادیر راست)
        ws['L14'] = inp.work_days
        ws['L15'] = r.base_salary
        ws['L16'] = r.housing_total
        ws['L17'] = r.grocery_total
        ws['L18'] = r.marriage_total
        ws['L20'] = r.children_total
        ws['L21'] = r.overtime_total
        ws['L22'] = r.friday_total
        ws['L23'] = r.night_total
        ws['L24'] = r.holiday_total
        ws['L25'] = r.shift_total
        ws['L26'] = r.severance_total
        ws['L27'] = r.bonus_total
        ws['L28'] = r.insurance_deduction
        ws['L29'] = r.unused_leave_amount
        ws['L30'] = inp.loan_deduction
        ws['L31'] = r.repair_wage

        # دیگر
        ws['B19'] = '2/5 روز در هر ماه'
        ws['P19'] = inp.leave_used
        ws['B20'] = inp.children_status or 'فاقد فرزند'
        ws['B25'] = inp.shift_type or 0
        ws['B28'] = '7% کل مبلغ  حق بیمه' if r.insurance_deduction > 0 else '—'
        ws['B29'] = inp.unused_leave

        # مبلغ کل
        ws['H32'] = r.grand_total

        # متن تعهد
        ws['B38'] = (
            f" بدین وسیله اینجانب: {inp.employee_name}"
            f" دارای شماره ملی: {inp.national_id}"
            f"  تصدیق و اعلام می دارد مبالغ مندرجه طبق شقوق جدول فوق را از کارفرمای خود"
            f"  {inp.employer_name}"
            f"  در طول مدت زمان مقرره به مبلغ {int(round(r.grand_total))} ریال"
            f" دریافت داشته که از  این حیث هـرگونه ادعاء برخلاف موارد فوق ،"
            f" در هرگونه مراجع قضایی و شبه قضایی غیر مسموع می باشد "
        )
        ws['B45'] = inp.employee_name

        # ═══════════════════════════════════════
        # ۲. فارسی‌سازی اعداد + فونت B Titr
        # ═══════════════════════════════════════

        # سلول‌هایی که سال هستن (بدون کاما)
        year_cells = {'F5', 'D10', 'N10', 'B12', 'K12'}

        for row in ws.iter_rows(min_row=3, max_row=47, min_col=1, max_col=22):
            for cell in row:
                if cell.value is None:
                    continue
                addr = cell.coordinate

                # B38 جداگانه هندل میشه
                if addr == 'B38':
                    text = str(cell.value)
                    # شماره ملی (10 رقم) → فارسی بدون کاما
                    text = re.sub(r'\b\d{10}\b', lambda m: _to_fa(m.group()), text)
                    # مبلغ (7-9 رقم) → فارسی با کاما
                    text = re.sub(r'\b\d{7,9}\b', lambda m: _fa_money(int(m.group())), text)
                    # سایر اعداد → فارسی
                    text = re.sub(r'\b\d+\b', lambda m: _to_fa(m.group()), text)
                    cell.value = text
                    _set_btitr(cell)
                    continue

                if isinstance(cell.value, (int, float)):
                    v = cell.value
                    if addr in year_cells:
                        cell.value = _to_fa(str(int(v)))
                    else:
                        if isinstance(v, float) and v != int(v):
                            # اعشاری — مثل ساعت کارکرد
                            cell.value = _fa_money(v)
                        else:
                            cell.value = _fa_money(v) if abs(v) >= 1000 else _fa_int(v)
                    _set_btitr(cell)

                elif isinstance(cell.value, str):
                    v = cell.value
                    if cell.number_format == '@' and v.strip().isdigit():
                        cell.value = _to_fa(v)
                    elif v.strip().startswith('7%'):
                        cell.value = _to_fa(v)
                    elif '2/5' in v:
                        cell.value = _to_fa(v)
                    # دو نقطه: از آخر بیار اول
                    if ':' in v and addr != 'B38':
                        stripped = v.replace(' : ', ' ').replace(':', '').strip()
                        cell.value = ': ' + stripped
                    if cell.font and cell.font.name in ('B Yekan', 'Calibri', 'B Nazanin'):
                        _set_btitr(cell)

        # ═══════════════════════════════════════
        # ۳. تنظیمات صفحه
        # ═══════════════════════════════════════

        ws.print_area = 'A3:V47'
        for sheet_name in wb.sheetnames:
            if sheet_name != 't,sahih(ساعت)':
                wb[sheet_name].sheet_state = 'hidden'

        ws.sheet_properties = WorksheetProperties()
        ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
        ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
        ws.page_setup.orientation = 'portrait'
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 1
        ws.page_margins = PageMargins(
            left=0.591, right=0.059,
            top=0.15, bottom=0.15,
            header=0, footer=0,
        )

        # ذخیره (بدون VBA — xlsx)
        xlsx_out = os.path.join(tmpdir, 'settlement.xlsx')
        wb.save(xlsx_out)

        # ═══════════════════════════════════════
        # ۴. تبدیل به PDF با LibreOffice
        # ═══════════════════════════════════════

        subprocess.run(
            ['libreoffice', '--headless', '--convert-to', 'pdf',
             '--outdir', tmpdir, xlsx_out],
            capture_output=True, timeout=120,
        )

        pdf_path = os.path.join(tmpdir, 'settlement.pdf')
        if not os.path.exists(pdf_path):
            raise RuntimeError("خطا در تبدیل PDF — LibreOffice نصب نیست یا خطا داد.")

        # استخراج صفحه ساعت
        from pypdf import PdfReader, PdfWriter
        import pdfplumber

        reader = PdfReader(pdf_path)
        target_page = None

        for i in range(len(reader.pages)):
            with pdfplumber.open(pdf_path) as pdf:
                words = pdf.pages[i].extract_words()
                text = ' '.join([w['text'] for w in words[:15]])
                if _to_fa(str(inp.year)) in text or inp.employee_name in text:
                    target_page = i
                    break

        if target_page is None:
            target_page = 0

        # وسط‌چین کردن
        with pdfplumber.open(pdf_path) as pdf:
            page = pdf.pages[target_page]
            words = page.extract_words()
            rects = page.rects
            all_x0 = [w['x0'] for w in words] + [rx['x0'] for rx in rects]
            all_x1 = [w['x1'] for w in words] + [rx['x1'] for rx in rects]

            if all_x0 and all_x1:
                left_gap = min(all_x0)
                right_gap = page.width - max(all_x1)
                shift = (right_gap - left_gap) / 2
            else:
                shift = 0

        writer = PdfWriter()
        p = reader.pages[target_page]
        if abs(shift) > 1:
            from pypdf import Transformation
            p.add_transformation(Transformation().translate(tx=shift, ty=0))
        writer.add_page(p)

        final_path = os.path.join(tmpdir, 'final.pdf')
        with open(final_path, 'wb') as f:
            writer.write(f)

        with open(final_path, 'rb') as f:
            pdf_bytes = f.read()

    buf = io.BytesIO(pdf_bytes)
    fname = f"فیش_تصفیه_{inp.employee_name}_{inp.year}_{inp.month_start:02d}.pdf"
    return buf, fname
