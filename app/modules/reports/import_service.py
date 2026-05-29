"""
سرویس Import اکسل — خواندن فایل، اعتبارسنجی، وارد کردن به DB.

جریان:
1. فایل اکسل پر شده → خوانده می‌شود
2. خلاصه به کاربر نشان داده می‌شود
3. داده‌ها وارد دیتابیس می‌شوند (با import_batch مشترک)
4. گزارش کامل + اکسل آپدیت‌شده (ستون وضعیت) ارسال می‌شود
5. کاربر می‌تواند بگوید «برگردون» → حذف بر اساس import_batch
"""
import io
import uuid
from datetime import datetime
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from sqlalchemy import select, delete
from sqlalchemy.ext.asyncio import AsyncSession

from app.database.models.business import Customer, Product, Employee, Expense, Invoice
from app.data.iran_geo import find_province
from app.utils.jalali import parse_jalali
from app.utils.id_generator import generate_display_id

SUCCESS_FILL = PatternFill("solid", start_color="DCFCE7")
ERROR_FILL = PatternFill("solid", start_color="FEE2E2")
STATUS_FONT = Font(name="Arial", size=10)
HEADER_FILL = PatternFill("solid", start_color="3B82F6")
HEADER_FONT = Font(bold=True, color="FFFFFF", name="Arial", size=11)
CENTER = Alignment(horizontal="center", vertical="center")
RIGHT = Alignment(horizontal="right", vertical="center")
BORDER = Border(*(Side(style="thin", color="E5E7EB"),) * 4)

# ─── نگاشت ستون‌ها به فیلدهای مدل ───
IMPORT_MAPS = {
    "customers": {
        "model": Customer,
        "required": ["name"],
        "columns": {
            "نام و نام خانوادگی": "name",
            "تلفن": "phone",
            "ایمیل": "email",
            "کد ملی": "national_id",
            "تاریخ تولد": "birth_date",
            "استان": "province",
            "شهر": "city",
            "آدرس": "address",
            "کد پستی": "postal_code",
            "سقف بدهی": "credit_limit",
            "کد": "code",
            "یادداشت": "note",
            "آیدی بله": "bale_id",
            "آیدی تلگرام": "telegram_id",
            "آیدی روبیکا": "rubika_id",
        },
        "date_fields": ["birth_date"],
        "numeric_fields": ["credit_limit"],
    },
    "products": {
        "model": Product,
        "required": ["name"],
        "columns": {
            "نام": "name",
            "دسته‌بندی": "category",
            "واحد": "unit",
            "قیمت خرید": "buy_price",
            "قیمت فروش": "sell_price",
            "موجودی": "stock",
            "حداقل موجودی": "min_stock",
            "تامین‌کننده": "supplier",
            "بارکد": "barcode",
            "کد": "code",
        },
        "date_fields": [],
        "numeric_fields": ["buy_price", "sell_price", "stock", "min_stock"],
    },
    "employees": {
        "model": Employee,
        "required": ["name"],
        "columns": {
            "نام و نام خانوادگی": "name",
            "کد ملی": "national_id",
            "تلفن": "phone",
            "تاریخ تولد": "birth_date",
            "نقش سازمانی": "role",
            "نحوه کار": "work_mode",
            "نوع قرارداد": "contract_type",
            "نوبت کاری": "shift_type",
            "کارکرد ماهانه (روز)": "monthly_work_days",
            "وضعیت تأهل": "marital_status",
            "تعداد فرزند": "children_count",
            "استان": "province",
            "شهر": "city",
            "آدرس": "address",
            "کد پستی": "postal_code",
            "حقوق پایه": "base_salary",
            "کسورات": "deductions",
            "شماره حساب": "bank_account",
            "شماره بیمه": "insurance_number",
            "مبلغ بیمه": "insurance_amount",
            "شروع بیمه": "insurance_start",
            "تاریخ استخدام": "hire_date",
            "پایان قرارداد": "contract_end",
            "مرخصی استحقاقی (روز)": "annual_leave",
            "کد پرسنلی": "code",
            "آیدی بله": "bale_id",
            "آیدی تلگرام": "telegram_id",
            "آیدی روبیکا": "rubika_id",
        },
        "date_fields": ["birth_date", "insurance_start", "hire_date", "contract_end"],
        "numeric_fields": ["children_count", "monthly_work_days", "base_salary",
                           "deductions", "annual_leave", "insurance_amount"],
    },
    "expenses": {
        "model": Expense,
        "required": ["title", "amount"],
        "columns": {
            "عنوان": "title",
            "مبلغ": "amount",
            "دسته‌بندی": "category",
            "نوع": "expense_type",
            "شخص": "person",
            "روش پرداخت": "payment_method",
            "تاریخ": "expense_date",
            "یادداشت": "note",
        },
        "date_fields": ["expense_date"],
        "numeric_fields": ["amount"],
    },
}

# ─── ذخیره آخرین batch هر کاربر برای بازگردانی ───
_last_import: dict[int, dict] = {}


def _detect_data_type(wb: Workbook) -> str | None:
    """تشخیص نوع داده از روی عنوان شیت یا ستون‌ها."""
    ws = wb.active
    title = (ws.title or "").strip()
    title_map = {"مشتریان": "customers", "کالاها": "products",
                 "کارمندان": "employees", "هزینه‌ها": "expenses"}
    if title in title_map:
        return title_map[title]

    # از روی ستون اول حدس بزن
    first_headers = []
    for col in range(1, min(ws.max_column or 1, 10) + 1):
        val = ws.cell(row=1, column=col).value
        if val:
            first_headers.append(str(val).strip())

    if "نام و نام خانوادگی" in first_headers:
        return "customers"
    if "عنوان" in first_headers and "مبلغ" in first_headers:
        return "expenses"
    if "حقوق پایه" in first_headers or "کد پرسنلی" in first_headers:
        return "employees"
    if "قیمت خرید" in first_headers or "قیمت فروش" in first_headers:
        return "products"

    return None


def _read_rows(wb: Workbook, mapping: dict) -> tuple[list[dict], list[str]]:
    """ردیف‌های اکسل را با mapping به dict تبدیل می‌کند."""
    ws = wb.active
    max_col = ws.max_column or 1

    # پیدا کردن mapping ستون‌ها
    header_map = {}  # col_index -> field_name
    for col in range(1, max_col + 1):
        header_val = ws.cell(row=1, column=col).value
        if header_val:
            header_str = str(header_val).strip()
            if header_str in mapping["columns"]:
                header_map[col] = mapping["columns"][header_str]

    rows = []
    errors = []

    for row_idx in range(2, (ws.max_row or 1) + 1):
        # رد ردیف‌های خالی
        all_empty = all(
            ws.cell(row=row_idx, column=c).value is None
            for c in header_map
        )
        if all_empty:
            continue

        # رد ردیف راهنما (اگر مقدار اولین فیلد الزامی شامل «الزامی» باشد)
        first_req = mapping["required"][0]
        for c, f in header_map.items():
            if f == first_req:
                v = ws.cell(row=row_idx, column=c).value
                if v and "الزامی" in str(v):
                    all_empty = True
                    break
        if all_empty:
            continue

        record = {"_row": row_idx}
        row_errors = []

        for col, field in header_map.items():
            val = ws.cell(row=row_idx, column=col).value
            if val is not None:
                val = str(val).strip() if not isinstance(val, (int, float)) else val
            record[field] = val

        # بررسی فیلدهای الزامی
        for req in mapping["required"]:
            if not record.get(req):
                row_errors.append(f"فیلد «{req}» خالی است")

        # تبدیل تاریخ‌ها
        for df in mapping["date_fields"]:
            if record.get(df):
                parsed = parse_jalali(str(record[df]))
                if parsed:
                    record[df] = parsed
                else:
                    row_errors.append(f"تاریخ «{df}» نامعتبر: {record[df]}")
                    record[df] = None
            else:
                record[df] = None

        # تبدیل اعداد
        for nf in mapping["numeric_fields"]:
            if record.get(nf) is not None:
                try:
                    record[nf] = float(str(record[nf]).replace(",", "").replace("٬", ""))
                except (ValueError, TypeError):
                    row_errors.append(f"عدد «{nf}» نامعتبر: {record[nf]}")
                    record[nf] = 0
            else:
                record[nf] = 0

        record["_errors"] = row_errors
        if row_errors:
            errors.append(f"ردیف {row_idx}: " + " | ".join(row_errors))

        rows.append(record)

    return rows, errors


def _make_result_excel(wb: Workbook, rows: list[dict], results: list[dict]) -> io.BytesIO:
    """اکسل آپدیت‌شده با ستون وضعیت."""
    ws = wb.active
    status_col = (ws.max_column or 1) + 1

    # هدر وضعیت
    c = ws.cell(row=1, column=status_col, value="وضعیت")
    c.fill = HEADER_FILL
    c.font = HEADER_FONT
    c.alignment = CENTER
    c.border = BORDER
    ws.column_dimensions[c.column_letter].width = 30

    for row_data, result in zip(rows, results):
        row_idx = row_data["_row"]
        cell = ws.cell(row=row_idx, column=status_col)
        cell.font = STATUS_FONT
        cell.alignment = RIGHT
        cell.border = BORDER
        if result["ok"]:
            cell.value = f"✅ {result['msg']}"
            cell.fill = SUCCESS_FILL
        else:
            cell.value = f"❌ {result['msg']}"
            cell.fill = ERROR_FILL

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


async def preview_import(file_bytes: bytes) -> tuple[str, str | None]:
    """
    پیش‌نمایش فایل اکسل — بدون ذخیره در DB.
    خروجی: (پیام خلاصه, نوع داده)
    """
    try:
        wb = load_workbook(io.BytesIO(file_bytes))
    except Exception:
        return "⚠️ فایل اکسل قابل خواندن نیست.", None

    data_type = _detect_data_type(wb)
    if not data_type:
        return "⚠️ نوع داده‌ی فایل رو نتونستم تشخیص بدم. لطفاً از template استفاده کن.", None

    mapping = IMPORT_MAPS[data_type]
    rows, errors = _read_rows(wb, mapping)

    if not rows:
        return "⚠️ فایل خالیه، هیچ ردیفی پیدا نشد.", None

    labels = {"customers": "مشتری", "products": "کالا",
              "employees": "کارمند", "expenses": "هزینه"}
    label = labels.get(data_type, data_type)
    valid = len(rows) - len(errors)

    msg = f"📋 {len(rows)} {label} پیدا کردم"
    if errors:
        msg += f" ({valid} سالم، {len(errors)} خطادار)"
    msg += ". وارد می‌کنم..."

    return msg, data_type


async def do_import(session: AsyncSession, tenant_id: int, user_id: int,
                    file_bytes: bytes) -> tuple[str, io.BytesIO | None, str | None]:
    """
    اجرای import واقعی.
    خروجی: (گزارش متنی, بافر اکسل آپدیت‌شده, نام فایل)
    """
    try:
        wb = load_workbook(io.BytesIO(file_bytes))
    except Exception:
        return "⚠️ فایل اکسل قابل خواندن نیست.", None, None

    data_type = _detect_data_type(wb)
    if not data_type:
        return "⚠️ نوع فایل تشخیص داده نشد.", None, None

    mapping = IMPORT_MAPS[data_type]
    model_class = mapping["model"]
    rows, _ = _read_rows(wb, mapping)

    if not rows:
        return "⚠️ فایل خالیه.", None, None

    batch_id = f"imp_{uuid.uuid4().hex[:8]}"
    results = []
    success_count = 0
    error_count = 0

    for row in rows:
        if row["_errors"]:
            results.append({"ok": False, "msg": " | ".join(row["_errors"])})
            error_count += 1
            continue

        # ساخت رکورد
        try:
            fields = {k: v for k, v in row.items() if not k.startswith("_") and v is not None}
            fields["tenant_id"] = tenant_id
            fields["import_batch"] = batch_id

            # شهر → استان
            if "city" in fields and fields["city"]:
                prov = find_province(fields["city"])
                if prov:
                    fields["province"] = prov

            # تولید آیدی
            if hasattr(model_class, "display_id"):
                did = await generate_display_id(session, tenant_id, model_class.__tablename__, model_class)
                fields["display_id"] = did

            obj = model_class(**fields)
            session.add(obj)
            await session.flush()

            did_str = getattr(obj, "display_id", "") or ""
            results.append({"ok": True, "msg": f"ثبت شد {did_str}"})
            success_count += 1
        except Exception as e:
            results.append({"ok": False, "msg": f"خطا: {str(e)[:60]}"})
            error_count += 1

    await session.commit()

    # ذخیره اطلاعات batch برای بازگردانی
    _last_import[user_id] = {
        "batch_id": batch_id,
        "data_type": data_type,
        "count": success_count,
        "tenant_id": tenant_id,
    }

    # ساخت اکسل آپدیت‌شده
    result_buf = _make_result_excel(wb, rows, results)

    labels = {"customers": "مشتریان", "products": "کالاها",
              "employees": "کارمندان", "expenses": "هزینه‌ها"}
    label = labels.get(data_type, data_type)
    ts = datetime.now().strftime("%Y%m%d_%H%M")
    filename = f"گزارش_import_{label}_{ts}.xlsx"

    # گزارش متنی
    report = f"📊 نتیجه import {label}:\n"
    report += f"✅ موفق: {success_count}\n"
    if error_count:
        report += f"❌ خطا: {error_count}\n"
    report += f"\nفایل آپدیت‌شده رو می‌فرستم."
    if success_count > 0:
        report += f"\n\n🔄 اگه پشیمون شدی بگو «برگردون» تا همه رو حذف کنم."

    return report, result_buf, filename


async def rollback_last_import(session: AsyncSession, user_id: int) -> str:
    """بازگردانی آخرین import."""
    info = _last_import.pop(user_id, None)
    if not info:
        return "⚠️ هیچ import اخیری برای بازگردانی وجود نداره."

    batch_id = info["batch_id"]
    data_type = info["data_type"]
    model_class = IMPORT_MAPS[data_type]["model"]

    result = await session.execute(
        delete(model_class).where(
            model_class.tenant_id == info["tenant_id"],
            model_class.import_batch == batch_id,
        )
    )
    await session.commit()

    labels = {"customers": "مشتریان", "products": "کالاها",
              "employees": "کارمندان", "expenses": "هزینه‌ها"}
    return f"🔄 {result.rowcount} {labels.get(data_type, '')} که import شده بود حذف شد. سیستم به حالت قبل برگشت."
