"""
سرویس خروجی اکسل — نسخه ۲.
- Export/Template/Import برای ۵ نوع
- اکسل زمان کار هر کارمند جدا
- zip وقتی بیش از ۱ فایل باشه
- نام‌گذاری دقیق فارسی
"""
import io
import zipfile
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession
from app.database.models.business import (
    Customer, Expense, Product, Employee, Invoice, InvoiceItem, WorkLog,
)
from app.utils.jalali import to_jalali_str
from app.utils.normalizer import format_amount

HEADER_FILL = PatternFill("solid", start_color="3B82F6")
HEADER_FONT = Font(bold=True, color="FFFFFF", name="Arial", size=11)
CELL_FONT = Font(name="Arial", size=10)
CENTER = Alignment(horizontal="center", vertical="center")
RIGHT = Alignment(horizontal="right", vertical="center")
BORDER = Border(*(Side(style="thin", color="E5E7EB"),) * 4)

# ─── تعریف ستون‌های هر بخش ───
SCHEMAS = {
    "customers": {
        "title": "مشتریان",
        "cols": [
            ("شناسه", 12, "display_id"), ("کد", 10, "code"),
            ("نام و نام خانوادگی", 25, "name"),
            ("تلفن", 16, "phone"), ("ایمیل", 20, "email"),
            ("کد ملی", 14, "national_id"), ("تاریخ تولد", 14, "birth_date$date"),
            ("استان", 14, "province"), ("شهر", 14, "city"),
            ("آدرس", 30, "address"), ("کد پستی", 14, "postal_code"),
            ("مانده حساب", 16, "balance$num"),
            ("سقف بدهی", 16, "credit_limit$num"), ("کل خرید", 16, "total_purchase$num"),
            ("یادداشت", 25, "note"),
            ("آیدی بله", 14, "bale_id"), ("آیدی تلگرام", 14, "telegram_id"),
            ("آیدی روبیکا", 14, "rubika_id"),
        ],
    },
    "products": {
        "title": "کالاها",
        "cols": [
            ("شناسه", 12, "display_id"), ("کد", 10, "code"),
            ("بارکد", 16, "barcode"), ("نام", 25, "name"),
            ("دسته‌بندی", 16, "category"), ("واحد", 10, "unit"),
            ("قیمت خرید", 16, "buy_price$num"), ("قیمت فروش", 16, "sell_price$num"),
            ("تخفیف", 12, "discount$num"), ("موجودی", 12, "stock"),
            ("حداقل موجودی", 14, "min_stock"), ("تامین‌کننده", 20, "supplier"),
        ],
    },
    "employees": {
        "title": "کارمندان",
        "cols": [
            ("شناسه", 12, "display_id"), ("کد پرسنلی", 12, "code"),
            ("نام و نام خانوادگی", 25, "name"),
            ("کد ملی", 14, "national_id"), ("تلفن", 16, "phone"),
            ("تاریخ تولد", 14, "birth_date$date"),
            ("نقش سازمانی", 18, "role"),
            ("نحوه کار", 14, "work_mode"), ("نوع قرارداد", 14, "contract_type"),
            ("نوبت کاری", 14, "shift_type"),
            ("کارکرد ماهانه (روز)", 16, "monthly_work_days"),
            ("وضعیت تأهل", 12, "marital_status"),
            ("تعداد فرزند", 10, "children_count"),
            ("استان", 14, "province"), ("شهر", 14, "city"),
            ("آدرس", 25, "address"), ("کد پستی", 14, "postal_code"),
            ("حقوق پایه", 16, "base_salary$num"), ("کسورات", 14, "deductions$num"),
            ("شماره حساب", 20, "bank_account"),
            ("شماره بیمه", 16, "insurance_number"), ("مبلغ بیمه", 14, "insurance_amount$num"),
            ("شروع بیمه", 14, "insurance_start$date"),
            ("تاریخ استخدام", 14, "hire_date$date"),
            ("پایان قرارداد", 14, "contract_end$date"),
            ("مرخصی استحقاقی (روز)", 16, "annual_leave"),
            ("آیدی بله", 14, "bale_id"), ("آیدی تلگرام", 14, "telegram_id"),
            ("آیدی روبیکا", 14, "rubika_id"),
        ],
    },
    "expenses": {
        "title": "هزینه‌ها",
        "cols": [
            ("عنوان", 25, "title"), ("مبلغ", 16, "amount$num"),
            ("دسته‌بندی", 16, "category"), ("نوع", 14, "expense_type"),
            ("شخص", 16, "person"), ("روش پرداخت", 14, "payment_method"),
            ("شماره مرجع", 14, "ref_number"), ("تاریخ", 14, "expense_date$date"),
            ("یادداشت", 25, "note"),
        ],
    },
    "invoices": {
        "title": "فاکتورها",
        "cols": [
            ("شناسه فاکتور", 12, "display_id"), ("شماره فاکتور", 14, "number"),
            ("شناسه مشتری", 14, "customer_id"),
            ("نام مشتری", 25, "customer_name"),
            ("جمع کل اقلام", 16, "total$num"), ("تخفیف کل", 14, "discount$num"),
            ("مالیات (9%)", 14, "tax$num"), ("مبلغ نهایی", 16, "final_amount$num"),
            ("پرداخت شده", 16, "paid$num"), ("مانده بدهی", 16, "remaining_debt$num"),
            ("وضعیت", 14, "status$invstatus"),
            ("روش پرداخت", 14, "payment_method"),
            ("تاریخ صدور", 14, "invoice_date$date"),
            ("یادداشت", 25, "note"),
        ],
    },
    "work_logs": {
        "title": "گزارش کار",
        "cols": [
            ("شناسه کارمند", 14, "employee_id"),
            ("نام کارمند", 20, "employee_name"),
            ("نحوه کار", 14, "work_mode"),
            ("کارکرد ساعتی", 14, "work_hours"),
            ("اضافه‌کاری (ساعت)", 14, "overtime_hours"),
            ("تعطیل‌کاری", 12, "is_holiday_work$bool"),
            ("شب‌کاری (ساعت)", 14, "night_hours"),
            ("جمعه‌کاری (روز)", 14, "friday_work"),
            ("مرخصی استفاده‌نشده (روز)", 18, "unused_leave"),
            ("مزد ترمیمی", 16, "repair_wage$num"),
            ("یادداشت", 25, "note"),
        ],
    },
}

MODELS = {
    "customers": Customer, "products": Product, "employees": Employee,
    "expenses": Expense, "invoices": Invoice,
}


def _setup_sheet(ws, schema):
    for i, (title, width, _) in enumerate(schema["cols"], start=1):
        c = ws.cell(row=1, column=i, value=title)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = CENTER
        c.border = BORDER
        ws.column_dimensions[c.column_letter].width = width
    ws.sheet_view.rightToLeft = True
    ws.freeze_panes = "A2"


_INV_STATUS_FA = {"draft": "پیش‌فاکتور", "confirmed": "تأیید‌شده",
                  "paid": "پرداخت شده", "installment": "اقساطی",
                  "cancelled": "لغو شده"}


def _cell_value(obj, field_spec):
    field = field_spec.split("$")[0]
    is_date = "$date" in field_spec
    is_num = "$num" in field_spec
    is_bool = "$bool" in field_spec
    is_invstatus = "$invstatus" in field_spec
    val = getattr(obj, field, None)
    if is_date:
        return to_jalali_str(val)
    if is_num:
        return float(val or 0)
    if is_bool:
        return "بله" if val else "خیر"
    if is_invstatus:
        return _INV_STATUS_FA.get(val, val or "")
    return val if val is not None else ""


def _apply_num_formats(ws, schema, n_rows):
    for col_idx, (_, _, spec) in enumerate(schema["cols"], start=1):
        if "$num" in spec:
            for r in range(2, n_rows + 2):
                ws.cell(row=r, column=col_idx).number_format = "#,##0"


def _finalize(wb):
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


async def export_data(session: AsyncSession, tenant_id: int, data_type: str):
    """خروجی اکسل داده‌ی واقعی."""
    schema = SCHEMAS[data_type]
    model = MODELS[data_type]
    rows_obj = (await session.scalars(
        select(model).where(model.tenant_id == tenant_id)
    )).all()

    wb = Workbook()
    ws = wb.active
    ws.title = schema["title"]
    _setup_sheet(ws, schema)

    for r_idx, obj in enumerate(rows_obj, start=2):
        for c_idx, (_, _, spec) in enumerate(schema["cols"], start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=_cell_value(obj, spec))
            cell.font = CELL_FONT
            cell.alignment = RIGHT
            cell.border = BORDER

    _apply_num_formats(ws, schema, len(rows_obj))
    ts = datetime.now().strftime("%Y%m%d_%H%M")
    filename = f"{schema['title']}_{ts}.xlsx"
    return _finalize(wb), filename


async def make_template(data_type: str):
    """اکسل نمونه‌ی خالی."""
    schema = SCHEMAS[data_type]
    wb = Workbook()
    ws = wb.active
    ws.title = schema["title"]
    _setup_sheet(ws, schema)

    hint = {"name": "(الزامی)", "title": "(الزامی)", "amount": "(الزامی)"}
    for c_idx, (_, _, spec) in enumerate(schema["cols"], start=1):
        field = spec.split("$")[0]
        cell = ws.cell(row=2, column=c_idx, value=hint.get(field, ""))
        cell.font = Font(name="Arial", size=9, italic=True, color="9CA3AF")
        cell.alignment = RIGHT
    filename = f"نمونه_{schema['title']}.xlsx"
    return _finalize(wb), filename


async def export_work_logs(session: AsyncSession, tenant_id: int,
                           employee_name: str = None):
    """
    خروجی اکسل گزارش کار کارمند(ان).
    اگر employee_name خالی باشه → همه کارمندان → zip.
    """
    query = select(Employee).where(Employee.tenant_id == tenant_id)
    if employee_name:
        query = query.where(Employee.name == employee_name)
    query = query.order_by(Employee.name)

    employees = (await session.scalars(query)).all()
    if not employees:
        return None, None, "کارمندی پیدا نشد."

    schema = SCHEMAS["work_logs"]
    files = []

    for emp in employees:
        logs = (await session.scalars(
            select(WorkLog).where(
                WorkLog.tenant_id == tenant_id,
                WorkLog.employee_id == emp.id,
            ).order_by(WorkLog.work_date)
        )).all()

        wb = Workbook()
        ws = wb.active
        ws.title = f"گزارش کار {emp.name}"

        # بنر اطلاعات کارمند (ردیف ۱)
        n_cols = len(schema["cols"])
        from openpyxl.utils import get_column_letter as _gcl
        ws.merge_cells(f"A1:{_gcl(n_cols)}1")
        banner = ws.cell(row=1, column=1,
                         value=f"گزارش کار: {emp.name}  |  شناسه: {emp.display_id or '—'}"
                               f"  |  کد پرسنلی: {emp.code or '—'}  |  نقش: {emp.role or '—'}")
        banner.font = Font(bold=True, name="Arial", size=11, color="1F2937")
        banner.alignment = RIGHT
        banner.fill = PatternFill("solid", start_color="EFF6FF")

        # هدر جدول (ردیف ۲)
        for i, (title, width, _) in enumerate(schema["cols"], start=1):
            c = ws.cell(row=2, column=i, value=title)
            c.fill = HEADER_FILL
            c.font = HEADER_FONT
            c.alignment = CENTER
            c.border = BORDER
            ws.column_dimensions[c.column_letter].width = width
        ws.sheet_view.rightToLeft = True
        ws.freeze_panes = "A3"

        for r_idx, log in enumerate(logs, start=3):
            for c_idx, (_, _, spec) in enumerate(schema["cols"], start=1):
                cell = ws.cell(row=r_idx, column=c_idx, value=_cell_value(log, spec))
                cell.font = CELL_FONT
                cell.alignment = RIGHT
                cell.border = BORDER

        # فرمت عددی (ردیف‌ها از ۳ شروع می‌شن)
        for col_idx, (_, _, spec) in enumerate(schema["cols"], start=1):
            if "$num" in spec:
                for r in range(3, len(logs) + 3):
                    ws.cell(row=r, column=col_idx).number_format = "#,##0"

        safe_name = emp.name.replace(" ", "_")
        did = emp.display_id or ""
        fname = f"گزارش‌کار_{safe_name}_{did}.xlsx"
        files.append((fname, _finalize(wb)))

    if len(files) == 1:
        return files[0][1], files[0][0], None

    # بیش از ۱ فایل → zip
    zip_buf = io.BytesIO()
    with zipfile.ZipFile(zip_buf, "w", zipfile.ZIP_DEFLATED) as zf:
        for fname, buf in files:
            zf.writestr(fname, buf.read())
    zip_buf.seek(0)
    ts = datetime.now().strftime("%Y%m%d_%H%M")
    return zip_buf, f"گزارش‌کار_همه‌کارمندان_{ts}.zip", None


async def make_work_log_template(employee_name: str = None):
    """اکسل نمونه گزارش کار."""
    schema = SCHEMAS["work_logs"]
    wb = Workbook()
    ws = wb.active
    ws.title = "گزارش کار"
    _setup_sheet(ws, schema)

    hints = {
        "work_date": "۱۴۰۵/۰۳/۰۱",
        "work_hours": "8",
        "overtime_hours": "2",
        "is_holiday_work": "بله/خیر",
        "night_hours": "0",
        "unused_leave": "0",
        "friday_work": "0",
        "repair_wage": "0",
    }
    for c_idx, (_, _, spec) in enumerate(schema["cols"], start=1):
        field = spec.split("$")[0]
        cell = ws.cell(row=2, column=c_idx, value=hints.get(field, ""))
        cell.font = Font(name="Arial", size=9, italic=True, color="9CA3AF")
        cell.alignment = RIGHT

    safe = employee_name.replace(" ", "_") if employee_name else "کارمند"
    return _finalize(wb), f"نمونه_گزارش‌کار_{safe}.xlsx"


# ─── نگاشت‌ها برای dispatcher ───
async def export_customers(s, t):
    return await export_data(s, t, "customers")


async def export_expenses(s, t):
    return await export_data(s, t, "expenses")


async def export_products(s, t):
    return await export_data(s, t, "products")


async def export_employees(s, t):
    return await export_data(s, t, "employees")


async def export_invoices(s, t):
    return await export_data(s, t, "invoices")


async def export_tenant(s, t):
    """اکسل اطلاعات کارفرما — حتی اگر ۱ ردیف باشه، با آیدی BIZ-001 اول."""
    from app.database.models.tenant import Tenant
    tenant = await s.get(Tenant, t)

    wb = Workbook()
    ws = wb.active
    ws.title = "اطلاعات فروشگاه"
    ws.sheet_view.rightToLeft = True

    cols = [
        ("شناسه", 12, "display_id"), ("نام فروشگاه", 25, "name"),
        ("تلفن", 16, "phone"), ("شهر", 14, "city"), ("استان", 14, "province"),
        ("آدرس", 35, "address"), ("شماره کارت", 22, "card_number"),
        ("شبا", 28, "sheba"), ("صاحب حساب", 22, "account_holder"),
        ("درصد مالیات", 12, "default_tax_percent"),
    ]
    for i, (title, width, _) in enumerate(cols, start=1):
        c = ws.cell(row=1, column=i, value=title)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = CENTER
        c.border = BORDER
        ws.column_dimensions[c.column_letter].width = width
    ws.freeze_panes = "A2"

    if tenant:
        for c_idx, (_, _, field) in enumerate(cols, start=1):
            val = getattr(tenant, field, None)
            cell = ws.cell(row=2, column=c_idx, value=val if val is not None else "")
            cell.font = CELL_FONT
            cell.alignment = RIGHT
            cell.border = BORDER

    ts = datetime.now().strftime("%Y%m%d_%H%M")
    return _finalize(wb), f"اطلاعات_فروشگاه_{ts}.xlsx"


EXPORTERS = {
    "customers": export_customers, "expenses": export_expenses,
    "products": export_products, "employees": export_employees,
    "invoices": export_invoices, "tenant": export_tenant,
}
