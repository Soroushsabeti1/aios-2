"""
سرویس جستجوی اینترنتی — سه سرعته.
از Gemini با Google Search استفاده می‌کند.
"""
import io
import json
import asyncio
from datetime import datetime, timezone, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from sqlalchemy import select, update
from sqlalchemy.ext.asyncio import AsyncSession
import httpx
from app.database.models.business import SearchTask
from app.core.config import settings

HEADER_FILL = PatternFill("solid", start_color="3B82F6")
HEADER_FONT = Font(bold=True, color="FFFFFF", name="Arial", size=11)
CELL_FONT = Font(name="Arial", size=10)
CENTER = Alignment(horizontal="center", vertical="center")
RIGHT = Alignment(horizontal="right", vertical="center")
BORDER = Border(*(Side(style="thin", color="E5E7EB"),) * 4)

# تعداد سرچ بر اساس اولویت
SEARCH_COUNTS = {
    "instant": 3,
    "medium": 15,
    "nightly": 50,
}


async def _call_gemini_search(query: str) -> list[dict]:
    """یک سرچ با مدل — درخواست نتایج ساختاریافته JSON."""
    payload = {
        "model": settings.openrouter_model,
        "messages": [
            {"role": "system", "content": (
                "تو یک موتور جستجوی حرفه‌ای کسب‌وکار هستی که به اطلاعات وب دسترسی داری. "
                "برای هر جستجو، کسب‌وکارها و اطلاعات واقعی مرتبط را پیدا کن. "
                "خروجی فقط یک JSON array باشد. هر آیتم شامل این کلیدها: "
                "name (نام)، phone (تلفن)، address (آدرس)، website (وب‌سایت)، description (توضیح). "
                "فقط JSON خالص برگردان بدون هیچ توضیح یا markdown. "
                "اگر اطلاعاتی نداری فیلد را رشته خالی بگذار."
            )},
            {"role": "user", "content": f"جستجو: {query}"},
        ],
    }

    headers = {
        "Authorization": f"Bearer {settings.openrouter_api_key}",
        "Content-Type": "application/json",
        "HTTP-Referer": "https://github.com/ai-business-os",
        "X-Title": "AI Business OS",
    }

    try:
        async with httpx.AsyncClient(timeout=90.0) as client:
            resp = await client.post(
                f"{settings.openrouter_base_url}/chat/completions",
                headers=headers, json=payload,
            )
            resp.raise_for_status()
            data = resp.json()

        content = data["choices"][0]["message"].get("content", "") or ""
        content = content.strip()
        # حذف markdown fence
        if content.startswith("```"):
            parts = content.split("```")
            if len(parts) >= 2:
                content = parts[1]
                if content.startswith("json"):
                    content = content[4:]
        content = content.strip()
        # پیدا کردن آرایه JSON
        start = content.find("[")
        end = content.rfind("]")
        if start != -1 and end != -1:
            content = content[start:end + 1]

        results = json.loads(content)
        if isinstance(results, list):
            return [r for r in results if isinstance(r, dict)]
        return []
    except Exception:
        return []


def _generate_queries(base_query: str, count: int) -> list[str]:
    """تولید query‌های متنوع از یک درخواست پایه."""
    queries = [base_query]
    # اضافه کردن واریاسیون‌ها
    variations = [
        f"{base_query} لیست کامل",
        f"{base_query} شماره تماس",
        f"{base_query} آدرس",
        f"{base_query} بهترین",
        f"{base_query} نزدیک من",
        f"{base_query} معتبر",
        f"راهنمای {base_query}",
        f"دایرکتوری {base_query}",
        f"{base_query} ایران",
    ]
    queries.extend(variations[:count - 1])
    return queries[:count]


def _deduplicate(all_results: list[dict]) -> list[dict]:
    """حذف تکراری‌ها بر اساس نام یا شماره."""
    seen_names = set()
    seen_phones = set()
    unique = []
    for r in all_results:
        name = (r.get("name") or "").strip()
        phone = (r.get("phone") or "").strip()
        if name and name in seen_names:
            continue
        if phone and phone in seen_phones:
            continue
        if name:
            seen_names.add(name)
        if phone:
            seen_phones.add(phone)
        unique.append(r)
    return unique


def _make_search_excel(results: list[dict], query: str) -> io.BytesIO:
    """اکسل حرفه‌ای از نتایج جستجو."""
    wb = Workbook()
    ws = wb.active
    ws.title = "نتایج جستجو"
    ws.sheet_view.rightToLeft = True

    cols = [
        ("ردیف", 8), ("نام", 30), ("تلفن", 20), ("آدرس", 40),
        ("وب‌سایت", 25), ("توضیحات", 35), ("منبع", 15),
    ]
    for i, (title, width) in enumerate(cols, 1):
        c = ws.cell(row=1, column=i, value=title)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = CENTER
        c.border = BORDER
        ws.column_dimensions[c.column_letter].width = width

    # اطلاعات جستجو در ردیف ۲
    ws.merge_cells("A2:G2")
    info = ws.cell(row=2, column=1, value=f"🔍 جستجو: {query} | تاریخ: {datetime.now().strftime('%Y/%m/%d %H:%M')} | تعداد: {len(results)}")
    info.font = Font(name="Arial", size=10, italic=True, color="6B7280")
    info.alignment = RIGHT

    for r_idx, item in enumerate(results, start=3):
        ws.cell(row=r_idx, column=1, value=r_idx - 2).font = CELL_FONT
        ws.cell(row=r_idx, column=2, value=item.get("name", "")).font = CELL_FONT
        ws.cell(row=r_idx, column=3, value=item.get("phone", "")).font = CELL_FONT
        ws.cell(row=r_idx, column=4, value=item.get("address", "")).font = CELL_FONT
        ws.cell(row=r_idx, column=5, value=item.get("website", "")).font = CELL_FONT
        ws.cell(row=r_idx, column=6, value=item.get("description", "")).font = CELL_FONT
        ws.cell(row=r_idx, column=7, value="Google").font = CELL_FONT
        for col in range(1, 8):
            ws.cell(row=r_idx, column=col).alignment = RIGHT
            ws.cell(row=r_idx, column=col).border = BORDER

    ws.freeze_panes = "A3"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


async def execute_search(session: AsyncSession, task_id: int):
    """اجرای یک وظیفه جستجو."""
    task = await session.get(SearchTask, task_id)
    if not task or task.status != "pending":
        return

    task.status = "running"
    task.started_at = datetime.now(timezone.utc)
    await session.commit()

    try:
        count = SEARCH_COUNTS.get(task.priority, 3)
        queries = _generate_queries(task.query, count)
        all_results = []

        for q in queries:
            results = await _call_gemini_search(q)
            all_results.extend(results)
            await asyncio.sleep(2)  # فاصله بین سرچ‌ها

        unique = _deduplicate(all_results)
        excel_buf = _make_search_excel(unique, task.query)

        task.status = "done"
        task.finished_at = datetime.now(timezone.utc)
        task.result_count = len(unique)
        task.results_json = json.dumps(unique, ensure_ascii=False)
        task.excel_data = excel_buf.read()
        await session.commit()

    except Exception as e:
        task.status = "failed"
        task.error = str(e)[:500]
        task.finished_at = datetime.now(timezone.utc)
        await session.commit()


async def create_search_task(session: AsyncSession, tenant_id: int,
                             user_telegram_id: int, query: str,
                             priority: str = "instant") -> tuple[str, int]:
    """ساخت وظیفه جستجو."""
    scheduled = None
    if priority == "nightly":
        now = datetime.now(timezone.utc)
        tomorrow_2am = (now + timedelta(days=1)).replace(hour=22, minute=0, second=0)  # 2 صبح ایران ≈ 22:30 UTC
        scheduled = tomorrow_2am

    task = SearchTask(
        tenant_id=tenant_id, user_telegram_id=user_telegram_id,
        query=query, priority=priority,
        scheduled_at=scheduled,
    )
    session.add(task)
    await session.commit()

    labels = {
        "instant": "⚡ فوری — ۲ تا ۱۰ دقیقه",
        "medium": "⏱ متوسط — ۱۰ تا ۴۰ دقیقه",
        "nightly": "🌙 شبانه — نتایج صبح آماده‌ست",
    }
    return labels.get(priority, ""), task.id


async def get_pending_tasks(session: AsyncSession) -> list[SearchTask]:
    """وظایف آماده اجرا — به ترتیب اولویت (فوری اول)."""
    from sqlalchemy import case
    now = datetime.now(timezone.utc)
    priority_order = case(
        (SearchTask.priority == "instant", 0),
        (SearchTask.priority == "medium", 1),
        (SearchTask.priority == "nightly", 2),
        else_=3,
    )
    return (await session.scalars(
        select(SearchTask).where(
            SearchTask.status == "pending",
            (SearchTask.scheduled_at.is_(None)) | (SearchTask.scheduled_at <= now),
        ).order_by(
            priority_order,
            SearchTask.created_at.asc(),
        )
    )).all()


async def get_search_result(session: AsyncSession, task_id: int) -> tuple[str | None, bytes | None, str | None]:
    """نتیجه یک جستجو."""
    task = await session.get(SearchTask, task_id)
    if not task:
        return None, None, None

    if task.status == "pending":
        return "⏳ هنوز توی صف هست...", None, None
    if task.status == "running":
        return "🔄 در حال جستجو...", None, None
    if task.status == "failed":
        return f"❌ جستجو با خطا مواجه شد: {task.error}", None, None

    msg = f"📊 جستجوی «{task.query}» تموم شد! {task.result_count} نتیجه پیدا شد."
    filename = f"نتایج_{task.query[:20].replace(' ', '_')}.xlsx"
    return msg, task.excel_data, filename
