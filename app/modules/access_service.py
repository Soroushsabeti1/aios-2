"""سرویس دسترسی‌ها — ۵ سطح با انواع محدودیت."""
import json
import io
from datetime import datetime, timezone, timedelta
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession
from app.database.models.business import AccessPermission


LEVEL_NAMES = {
    1: "عمومی", 2: "سمتی", 3: "تأیید کارفرما",
    4: "محرمانه", 5: "سفارشی",
}


async def grant_permission(session: AsyncSession, tenant_id: int,
                            name: str, level: int,
                            grantee_type: str, resource_type: str,
                            grantee_id: int = None, grantee_role: str = None,
                            resource_filter: dict = None,
                            resource_exclude: dict = None,
                            max_uses: int = None,
                            expires_hours: int = None,
                            condition: str = None) -> str:
    expires_at = None
    if expires_hours:
        expires_at = datetime.now(timezone.utc) + timedelta(hours=expires_hours)

    perm = AccessPermission(
        tenant_id=tenant_id, name=name, level=level,
        grantee_type=grantee_type, grantee_id=grantee_id,
        grantee_role=grantee_role, resource_type=resource_type,
        resource_filter=json.dumps(resource_filter, ensure_ascii=False) if resource_filter else None,
        resource_exclude=json.dumps(resource_exclude, ensure_ascii=False) if resource_exclude else None,
        max_uses=max_uses, expires_at=expires_at, condition=condition,
        is_active=True,
    )
    session.add(perm)
    await session.commit()
    return f"✅ دسترسی «{name}» (سطح {level}: {LEVEL_NAMES.get(level, '')}) ثبت شد."


async def check_permission(session: AsyncSession, tenant_id: int,
                            grantee_id: int, resource_type: str,
                            grantee_role: str = None) -> bool:
    """بررسی دسترسی — True=مجاز."""
    now = datetime.now(timezone.utc)
    perms = (await session.scalars(
        select(AccessPermission).where(
            AccessPermission.tenant_id == tenant_id,
            AccessPermission.resource_type == resource_type,
            AccessPermission.is_active == True,
        )
    )).all()

    for p in perms:
        # چک انقضا
        if p.expires_at:
            exp = p.expires_at
            if exp.tzinfo is None:
                exp = exp.replace(tzinfo=timezone.utc)
            if exp < now:
                continue
        # چک لیمیت
        if p.max_uses and p.use_count >= p.max_uses:
            continue
        # چک هدف
        if p.grantee_type == "all":
            return True
        if p.grantee_type == "person" and p.grantee_id == grantee_id:
            p.use_count += 1
            await session.commit()
            return True
        if p.grantee_type == "role" and p.grantee_role == grantee_role:
            return True

    return False


async def list_permissions(session: AsyncSession, tenant_id: int) -> str:
    perms = (await session.scalars(
        select(AccessPermission).where(
            AccessPermission.tenant_id == tenant_id,
            AccessPermission.is_active == True,
        ).order_by(AccessPermission.level)
    )).all()

    if not perms:
        return "هیچ دسترسی خاصی تعریف نشده."

    lines = ["🔐 دسترسی‌ها:"]
    for p in perms:
        level_name = LEVEL_NAMES.get(p.level, "")
        lines.append(f"• [{p.id}] {p.name} — سطح {p.level} ({level_name}) — {p.resource_type}")
        if p.grantee_role:
            lines.append(f"   برای: {p.grantee_role}")
        if p.max_uses:
            lines.append(f"   لیمیت: {p.use_count}/{p.max_uses}")
        if p.expires_at:
            from app.utils.jalali import to_jalali_str
            lines.append(f"   انقضا: {to_jalali_str(p.expires_at.date() if hasattr(p.expires_at, 'date') else p.expires_at)}")
    return "\n".join(lines)


async def revoke_permission(session: AsyncSession, tenant_id: int, perm_id: int) -> str:
    perm = await session.get(AccessPermission, perm_id)
    if not perm or perm.tenant_id != tenant_id:
        return "⚠️ دسترسی پیدا نشد."
    perm.is_active = False
    await session.commit()
    return f"🚫 دسترسی «{perm.name}» لغو شد."


async def export_permissions_excel(session: AsyncSession, tenant_id: int):
    """خروجی اکسل دسترسی‌ها — محرمانه، فقط کارفرما."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill

    perms = (await session.scalars(
        select(AccessPermission).where(AccessPermission.tenant_id == tenant_id)
    )).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "دسترسی‌ها"
    ws.sheet_view.rightToLeft = True

    headers = ["شناسه", "نام", "سطح", "نوع گیرنده", "نوع منبع",
               "لیمیت استفاده", "انقضا", "شرط", "وضعیت"]
    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=i, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="2B5F9E")

    for r, p in enumerate(perms, 2):
        ws.cell(row=r, column=1, value=p.id)
        ws.cell(row=r, column=2, value=p.name)
        ws.cell(row=r, column=3, value=f"{p.level} - {LEVEL_NAMES.get(p.level,'')}")
        ws.cell(row=r, column=4, value=p.grantee_role or p.grantee_type)
        ws.cell(row=r, column=5, value=p.resource_type)
        ws.cell(row=r, column=6, value=f"{p.use_count}/{p.max_uses}" if p.max_uses else "نامحدود")
        ws.cell(row=r, column=7, value=str(p.expires_at.date()) if p.expires_at else "بدون انقضا")
        ws.cell(row=r, column=8, value=p.condition or "—")
        ws.cell(row=r, column=9, value="فعال" if p.is_active else "غیرفعال")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf, "دسترسی‌ها.xlsx"
